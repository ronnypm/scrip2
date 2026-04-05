"""
Excel exporter for scraped book data.
"""
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Export scraped books to Excel file.
    """
    
    # Column headers in Spanish
    HEADERS = [
        "Título",
        "Autor",
        "Descripción",
        "Precio (S/)",
        "Envío Rápido",
        "URL",
        "Fecha Extracción"
    ]
    
    # Column keys (mapping to book dict)
    COLUMN_KEYS = [
        "titulo",
        "autor",
        "descripcion",
        "precio",
        "envio_rapido",
        "url",
        "fecha_extraccion"
    ]
    
    # Header styles
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Cell styles
    CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    def __init__(self):
        pass
    
    def export(self, books: List[Dict[str, Any]], output_dir: str = "./output") -> str:
        """
        Export books to Excel file.
        
        Args:
            books: List of book dictionaries.
            output_dir: Directory to save the Excel file.
        
        Returns:
            Path to the generated Excel file.
        
        Raises:
            IOError: If unable to write the file.
        """
        if not books:
            logger.warning("No hay libros para exportar")
            # Create empty file
            return self._create_empty_file(output_dir)
        
        # Create output directory if it doesn't exist
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        filename = f"buscalibre_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        logger.info(f"Generando Excel: {filepath}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Libros"
        
        # Write headers
        self._write_headers(ws)
        
        # Write data
        self._write_data(ws, books)
        
        # Adjust column widths
        self._adjust_column_widths(ws)
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save file
        wb.save(filepath)
        
        logger.info(f"Excel generado con {len(books)} libros")
        
        return filepath
    
    def _write_headers(self, ws):
        """Write header row with styling."""
        for col_num, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.BORDER
    
    def _write_data(self, ws, books: List[Dict[str, Any]]):
        """Write book data rows."""
        for row_num, book in enumerate(books, start=2):
            for col_num, key in enumerate(self.COLUMN_KEYS, start=1):
                value = book.get(key, "")
                
                # Format values
                if key == "precio":
                    value = self._format_price(value)
                elif key == "envio_rapido":
                    value = "Sí" if value else "No"
                elif key == "descripcion":
                    # Truncate long descriptions
                    value = str(value)[:2000] if value else ""
                
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = self.CELL_ALIGN
                cell.border = self.BORDER
    
    def _format_price(self, price) -> str:
        """Format price for display."""
        if price is None:
            return ""
        
        try:
            return f"S/ {float(price):.2f}"
        except (ValueError, TypeError):
            return str(price)
    
    def _adjust_column_widths(self, ws):
        """Adjust column widths based on content."""
        column_widths = {
            1: 40,  # Título
            2: 25,  # Autor
            3: 50,  # Descripción
            4: 12,  # Precio
            5: 12,  # Envío rápido
            6: 50,  # URL
            7: 20   # Fecha
        }
        
        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width
    
    def _create_empty_file(self, output_dir: str) -> str:
        """Create an empty Excel file when no books are extracted."""
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        filename = f"buscalibre_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Libros"
        
        # Write headers only
        self._write_headers(ws)
        
        wb.save(filepath)
        
        logger.info(f"Archivo Excel vacío creado: {filepath}")
        
        return filepath